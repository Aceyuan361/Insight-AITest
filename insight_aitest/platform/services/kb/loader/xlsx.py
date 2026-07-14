# -*- coding: utf-8 -*-
"""Excel 表格 Loader（.xlsx）。用 openpyxl 逐 sheet 提取每行单元格文本。"""

from __future__ import annotations

from pathlib import Path

from insight_aitest.platform.services.kb.loader import register_loader
from insight_aitest.platform.services.kb.loader.base import DocumentLoader
from insight_aitest.platform.services.kb.models import ParsedDocument


@register_loader(".xlsx")
class XlsxLoader(DocumentLoader):
    """Excel .xlsx 文档 Loader。逐 sheet 提取表头与各行，单元格用 ` | ` 拼接。"""

    def load(self, path: Path) -> ParsedDocument:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True, read_only=True)
        parts: list[str] = []

        for sheet in wb.worksheets:
            parts.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    parts.append(" | ".join(cells))

        wb.close()
        content = "\n".join(parts)
        return ParsedDocument(filename=path.name, content=content, meta={"format": "xlsx"})
